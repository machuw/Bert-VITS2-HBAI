import openpyxl
from multiprocessing import Pool
import sys
import os
import argparse
from tqdm import tqdm

root_dir_m = "../../datasets/Genshin/English/"

def find_file_path(root_dir, target_file_name):
    # 使用os.walk递归遍历root_dir及其所有子目录
    for dirpath, dirnames, filenames in os.walk(root_dir):
        # 检查当前目录中的每个文件名
        for filename in filenames:
            # 如果文件名与目标文件名匹配，则打印文件路径
            if filename == target_file_name:
                #print(os.path.join(dirpath, filename))
                return os.path.join(dirpath, filename)

def process_line(row):
    if not row[2].endswith('wav'):
        return
    target_file_name = row[2]
    lab_text = row[3]
    wav_file = find_file_path(root_dir_m, target_file_name)
    lab_file = wav_file.replace('.wav', '.lab')
    if os.path.exists(lab_file):
        return 
    with open(lab_file, "w", encoding="utf-8") as f:
        f.write(lab_text)

def main_mp(args):
    # 指定要开始搜索的根目录
    root_dir = args.root_dir
    
    # 加载工作簿
    workbook = openpyxl.load_workbook(args.in_dir)
    
    # 遍历工作簿中的所有工作表
    lines = []
    for sheet_name in tqdm(workbook.sheetnames):
        if sheet_name=='总览数据':
            continue
        sheet = workbook[sheet_name]
        lines.extend(sheet.iter_rows(values_only=True))

    # 遍历工作表中的所有行
    num_processes = args.num_processes
    with Pool(processes=num_processes) as pool:
        for _ in tqdm(pool.imap_unordered(process_line, lines), total=len(lines)):
            pass


def main(args):
    # 指定要开始搜索的根目录
    root_dir = args.root_dir
    
    # 加载工作簿
    workbook = openpyxl.load_workbook(args.in_dir)
    
    # 遍历工作簿中的所有工作表
    for sheet_name in tqdm(workbook.sheetnames):
        if sheet_name=='总览数据':
            continue
        sheet = workbook[sheet_name]

        # 遍历工作表中的所有行
        for row in tqdm(sheet.iter_rows(values_only=True)):
            if not row[2].endswith('wav'):
                continue
            target_file_name = row[2]
            lab_text = row[3]
            wav_file = find_file_path(root_dir, target_file_name)
            lab_file =  wav_file.replace('.wav', '.lab')
            if os.path.exists(lab_file):
                continue
            with open(lab_file, "w", encoding="utf-8") as f:
                f.write(lab_text)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--in_dir", type=str, default="/root/autodl-tmp/datasets/Genshin/原神4.0语音包对应文本（英）.xlsx", help="path to source dir"
    )
    #parser.add_argument(
    #    "--out_dir", type=str, default="./English", help="path to source dir"
    #)
    parser.add_argument(
        "--root_dir", type=str, default="../../datasets/Honkai/English/", help="path to target dir"
    )
    parser.add_argument("--num_processes", type=int, default=20)
    args = parser.parse_args()
    main_mp(args)

